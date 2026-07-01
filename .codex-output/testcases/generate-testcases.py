from copy import copy
from pathlib import Path
from openpyxl import load_workbook

OUT_DIR = Path(r"D:\桌面\大三下\小学期\JoyGather\.codex-output\testcases\outputs")
OUT_DIR.mkdir(parents=True, exist_ok=True)

MANUAL_TEMPLATE = Path(r"D:\Downloads\importTemplate-20260630094947.xlsx")
API_TEMPLATE = Path(r"D:\Downloads\Import-Testcase-Template-20260518195244.xlsx")

MANUAL_OUT = OUT_DIR / "JoyGather_华为云_手工测试用例_20260630.xlsx"
API_OUT = OUT_DIR / "JoyGather_华为云_API接口测试用例_20260630.xlsx"


def clone_row_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def steps(*pairs):
    flat = []
    for action, expected in pairs[:7]:
        flat.extend([action, expected])
    while len(flat) < 14:
        flat.extend(["", ""])
    return flat


manual_cases = [
    ("JG-AUTH-001", "身份管理", "邮箱注册成功并生成激活令牌", "Priority 1", "验证个人用户通过邮箱、昵称和密码注册。", "打开 /auth，网络正常。", steps(
        ("选择注册，输入未使用邮箱、昵称、8位以上密码和确认密码", "注册请求提交成功，页面提示账号待激活或展示激活入口"),
        ("查看返回的本地激活链接或激活 token", "系统生成 activationToken，账号状态为待激活"),
        ("使用该账号直接登录", "系统拒绝登录并提示请先激活账号"),
        ("访问激活链接后再次登录", "登录成功并跳转用户首页")
    )),
    ("JG-AUTH-002", "身份管理", "邮箱格式非法时注册失败", "Priority 1", "覆盖 RegisterRequest 邮箱校验异常。", "打开注册页。", steps(
        ("输入 abc、合法昵称和一致密码后提交", "表单或接口返回明确错误提示，不创建账号"),
        ("改为 user@example.com 后重新提交", "注册流程可继续")
    )),
    ("JG-AUTH-003", "身份管理", "两次密码不一致时注册失败", "Priority 1", "覆盖 confirmPassword 校验。", "打开注册页。", steps(
        ("输入合法邮箱，密码 12345678，确认密码 12345679", "系统阻止提交并提示两次密码不一致"),
        ("修正确认密码后提交", "注册成功进入待激活状态")
    )),
    ("JG-AUTH-004", "身份管理", "重复邮箱注册被拒绝", "Priority 1", "使用种子账号 demo@quju.cn 验证唯一性。", "数据库已导入 V2 种子数据。", steps(
        ("注册邮箱 demo@quju.cn", "接口返回邮箱已存在或同等含义错误"),
        ("检查用户列表", "未产生重复用户记录")
    )),
    ("JG-AUTH-005", "身份管理", "普通用户登录成功并保存 Token", "Priority 1", "验证 /auth/login 和前端 localStorage。", "存在 demo@quju.cn / 12345678。", steps(
        ("输入 demo@quju.cn 和 12345678 登录", "登录成功，返回 token 和用户信息"),
        ("查看首页或个人中心", "页面处于已登录状态，显示昵称小满"),
        ("刷新页面", "Token 有效期内无需重复登录")
    )),
    ("JG-AUTH-006", "身份管理", "登录密码错误时提示失败", "Priority 1", "覆盖登录异常路径。", "存在 demo@quju.cn。", steps(
        ("输入 demo@quju.cn 和错误密码登录", "系统提示邮箱或密码错误"),
        ("检查 localStorage", "不写入 quju:token 或 quju:session")
    )),
    ("JG-AUTH-007", "身份管理", "封禁用户登录被拦截", "Priority 1", "覆盖已封禁状态，种子用户 momo77@quju.cn。", "存在 momo77@quju.cn / 12345678。", steps(
        ("输入 momo77@quju.cn 和 12345678 登录", "登录失败，提示账号已被封禁"),
        ("查看提示信息", "展示封禁原因或期限，不进入首页")
    )),
    ("JG-AUTH-008", "身份管理", "退出登录清理本地会话", "Priority 2", "覆盖 /auth/logout 与 clearAuthStorage。", "普通用户已登录。", steps(
        ("点击退出登录", "调用 logout 接口并清理 token/session"),
        ("访问个人中心或报名功能", "系统要求重新登录")
    )),
    ("JG-PROFILE-001", "个人资料", "查看个人中心基础资料", "Priority 2", "验证 ProfileView 与 /auth/me。", "普通用户已登录。", steps(
        ("进入个人中心", "展示头像、昵称、城市、签名、兴趣标签等信息"),
        ("刷新页面", "资料仍从后端正确加载")
    )),
    ("JG-PROFILE-002", "个人资料", "编辑昵称与简介保存成功", "Priority 2", "覆盖 PUT /users/me。", "普通用户已登录。", steps(
        ("进入个人资料编辑区，修改昵称和简介", "保存按钮可点击"),
        ("点击保存", "接口返回成功，页面即时展示新资料"),
        ("重新打开个人中心", "修改内容仍然存在")
    )),
    ("JG-PROFILE-003", "个人资料", "头像图片上传成功", "Priority 2", "覆盖 /files/upload 上传头像。", "准备 jpg 或 png 图片，用户已登录。", steps(
        ("选择头像文件并上传", "上传成功，返回可访问 URL"),
        ("保存个人资料", "头像更新为新图片"),
        ("刷新页面", "新头像正常显示")
    )),
    ("JG-DISC-001", "活动发现", "首页最新活动按发布时间展示", "Priority 1", "覆盖 HomeView 最新信息流。", "后端已启动并有种子活动。", steps(
        ("打开用户首页", "活动卡片列表正常加载"),
        ("选择最新 Tab", "活动按发布时间或默认排序展示"),
        ("检查活动卡片", "标题、时间、地点、价格、人数状态完整")
    )),
    ("JG-DISC-002", "活动发现", "推荐活动列表可打开详情", "Priority 1", "覆盖 ActivityCard 到详情页跳转。", "首页活动列表已加载。", steps(
        ("点击任一活动卡片", "跳转到 /activities/:id 详情页"),
        ("查看详情信息", "展示简介、描述、组织者、地点、报名人数和安全提示")
    )),
    ("JG-DISC-003", "活动发现", "关键词搜索活动", "Priority 1", "覆盖 GET /activities?keyword。", "活动数据包含桌游、飞盘、Citywalk 等关键词。", steps(
        ("在搜索框输入 桌游 并提交", "列表仅展示与桌游相关活动"),
        ("清空关键词", "列表恢复展示更多活动")
    )),
    ("JG-DISC-004", "活动发现", "多选分类筛选活动", "Priority 2", "覆盖 category/categories 参数。", "活动列表页已打开。", steps(
        ("选择运动健身和户外运动分类", "列表展示匹配任一分类的活动"),
        ("取消全部分类", "筛选条件清空，列表恢复")
    )),
    ("JG-DISC-005", "活动发现", "按城市和城区筛选活动", "Priority 2", "覆盖 city/district 相关显示。", "活动数据包含杭州不同城区。", steps(
        ("选择杭州并筛选西湖区附近活动", "展示九溪、太子湾或西湖区活动"),
        ("切换其他城区", "列表随筛选条件刷新")
    )),
    ("JG-DISC-006", "活动发现", "按费用筛选免费活动", "Priority 2", "覆盖 fee 参数。", "存在 price 为 0 的活动。", steps(
        ("选择免费筛选", "仅展示免费或价格为 0 的活动"),
        ("选择全部费用", "收费活动重新出现")
    )),
    ("JG-DISC-007", "活动发现", "地图模式展示活动点位", "Priority 2", "覆盖 CityMap 组件。", "浏览器允许页面渲染地图区域。", steps(
        ("进入发现页并切换地图模式", "地图区域加载，活动点位可见"),
        ("点击点位", "展示对应活动摘要或卡片"),
        ("点击卡片进入详情", "跳转到对应活动详情")
    )),
    ("JG-DISC-008", "活动发现", "活动详情不存在时显示错误状态", "Priority 2", "覆盖 GET /activities/{id} 404。", "后端已启动。", steps(
        ("访问不存在的活动详情地址 /activities/not-exist", "页面提示活动不存在或加载失败"),
        ("返回列表页", "可继续浏览正常活动")
    )),
    ("JG-ACT-001", "活动创建", "基础活动创建成功并进入审核或发布", "Priority 1", "覆盖 POST /activities。", "普通用户已登录。", steps(
        ("进入创建活动页，填写标题、简介、分类、日期、时间、地点、人数上限", "必填字段均通过校验"),
        ("填写安全提示、年龄限制、报名字段和标签", "表单可继续提交"),
        ("点击提交发布", "接口返回新活动，状态为报名中或待审核"),
        ("在活动列表或后台审核查看", "新活动可被查询到")
    )),
    ("JG-ACT-002", "活动创建", "创建活动缺少标题时失败", "Priority 1", "覆盖 @NotBlank title。", "普通用户已登录。", steps(
        ("创建活动时留空标题，其余必填项填写完整", "系统阻止提交"),
        ("查看错误提示", "提示活动名称不能为空")
    )),
    ("JG-ACT-003", "活动创建", "创建活动人数小于2时失败", "Priority 1", "覆盖 @Min capacity。", "普通用户已登录。", steps(
        ("创建活动人数上限填写 1", "系统阻止提交或接口返回校验失败"),
        ("改为 2 后提交", "人数校验通过")
    )),
    ("JG-ACT-004", "活动创建", "地图选点自动填充地址坐标", "Priority 2", "覆盖 LocationPicker。", "普通用户已登录。", steps(
        ("在创建页打开地图选点", "显示搜索和地图点选能力"),
        ("搜索 西湖文化广场 并选择地点", "地址、经度、纬度自动回填"),
        ("提交活动", "活动详情展示选点地址")
    )),
    ("JG-ACT-005", "活动创建", "地图服务不可用时可手动输入地点", "Priority 2", "覆盖高德降级。", "模拟地图搜索失败或不配置高德 Key。", steps(
        ("在地点字段手动输入 天目里社区中心", "地点文本可保存"),
        ("提交活动", "活动创建不被地图服务失败阻断")
    )),
    ("JG-ACT-006", "活动创建", "保存草稿允许缺少必填字段", "Priority 1", "覆盖 POST /activities/drafts。", "普通用户已登录。", steps(
        ("创建活动只填写标题和简介", "保存草稿按钮可点击"),
        ("点击保存草稿", "草稿保存成功"),
        ("进入草稿列表", "草稿出现且不在首页公开列表展示")
    )),
    ("JG-ACT-007", "活动创建", "编辑草稿后提交审核", "Priority 1", "覆盖 PUT /activities/drafts/{id} 和 /submit。", "已有草稿。", steps(
        ("打开草稿继续编辑", "表单带出草稿内容"),
        ("补齐分类、日期、时间、地点、人数上限", "提交按钮可用"),
        ("点击提交", "草稿转为活动并进入发布或审核流程")
    )),
    ("JG-ACT-008", "活动创建", "删除草稿后不可再查看", "Priority 2", "覆盖 DELETE /activities/drafts/{id}。", "已有草稿。", steps(
        ("在草稿列表点击删除", "系统提示删除成功"),
        ("刷新草稿列表", "该草稿不再出现"),
        ("直接访问草稿提交接口", "系统返回无权限或不存在")
    )),
    ("JG-ACT-009", "活动创建", "克隆已发布活动为草稿", "Priority 2", "覆盖 POST /activities/{id}/clone。", "普通用户已登录，存在 act-001。", steps(
        ("在活动详情点击克隆", "系统创建新草稿"),
        ("查看草稿内容", "标题、简介、地点、人数设置被复制"),
        ("检查报名与签到数据", "报名记录、候补、评价不被复制")
    )),
    ("JG-ACT-010", "活动创建", "超过50人的活动进入人工审核", "Priority 1", "覆盖 ActivityService 审核分流。", "普通用户已登录。", steps(
        ("创建人数上限 68 的活动并提交", "提交成功但不直接公开"),
        ("管理员打开审核列表", "出现该活动审核任务，风险原因包含人数超过50或类似说明")
    )),
    ("JG-ACT-011", "活动创建", "风险文本活动进入人工审核", "Priority 1", "覆盖 AI 内容安全审核分流。", "普通用户已登录。", steps(
        ("创建简介包含高风险夜间/危险活动描述的活动", "活动提交成功"),
        ("查看后台审核列表", "该活动出现在待审核任务中，保留风险提示")
    )),
    ("JG-ACT-012", "活动创建", "重复提交不产生重复审核任务", "Priority 2", "覆盖 submitToken 幂等场景。", "创建页已填写完整活动信息。", steps(
        ("连续快速点击提交按钮两次", "前端只保留一次提交状态或后端只创建一个活动"),
        ("查看后台审核或活动列表", "不存在重复标题的重复任务")
    )),
    ("JG-REG-001", "报名候补", "活动报名成功", "Priority 1", "覆盖 POST /activities/{id}/registrations。", "普通用户已登录，活动 act-002 未满员。", steps(
        ("打开 act-002 详情页", "报名按钮可用"),
        ("填写真实姓名、手机号码等报名字段", "字段校验通过"),
        ("点击报名", "报名成功，状态显示已报名，人数增加")
    )),
    ("JG-REG-002", "报名候补", "重复报名同一活动被拦截", "Priority 1", "覆盖重复注册保护。", "用户已报名某活动。", steps(
        ("再次点击同一活动报名", "系统提示已报名或返回当前报名状态"),
        ("刷新我的报名状态", "仍只有一条报名记录")
    )),
    ("JG-REG-003", "报名候补", "满员活动加入候补", "Priority 1", "覆盖候补队列。", "活动 act-003 容量已满或构造满员活动。", steps(
        ("打开满员活动详情并点击报名", "系统提示加入候补"),
        ("查看我的报名状态", "该活动状态为候补中并显示队列位置")
    )),
    ("JG-REG-004", "报名候补", "取消报名后候补自动递补", "Priority 1", "覆盖 DELETE /registrations/me。", "活动存在已报名用户和候补用户。", steps(
        ("已报名用户取消报名", "取消成功，活动名额释放"),
        ("候补首位用户刷新报名状态", "状态从候补中变为已报名或可确认"),
        ("查看活动人数", "人数与状态同步更新")
    )),
    ("JG-REG-005", "报名候补", "确认候补递补名额", "Priority 2", "覆盖 /waitlist/confirm。", "用户收到候补递补资格。", steps(
        ("点击确认候补", "接口返回报名成功"),
        ("查看活动详情", "用户状态显示已报名")
    )),
    ("JG-CHK-001", "签到评价", "发起人生成活动签到二维码", "Priority 2", "覆盖 POST /activities/{id}/checkins/qr。", "活动发起人已登录，活动存在。", steps(
        ("进入本人发起的活动详情或签到管理", "显示生成二维码入口"),
        ("点击生成签到码", "返回 code、url 和过期时间"),
        ("刷新页面", "可继续展示有效签到信息")
    )),
    ("JG-CHK-002", "签到评价", "报名用户扫码签到成功", "Priority 1", "覆盖 POST /activities/checkins/scan。", "用户已报名，拥有有效签到码。", steps(
        ("进入签到页或扫描二维码", "系统读取签到 code"),
        ("提交签到", "签到成功，报名状态更新为已签到"),
        ("重复签到", "系统提示已签到或返回当前状态")
    )),
    ("JG-CHK-003", "签到评价", "无效签到码签到失败", "Priority 2", "覆盖扫码异常。", "普通用户已登录。", steps(
        ("在签到页输入无效 code", "系统拒绝签到"),
        ("查看提示", "展示二维码无效、过期或无权限等错误")
    )),
    ("JG-REV-001", "签到评价", "活动结束后发布活动总结", "Priority 2", "覆盖 /activities/{id}/summaries。", "活动发起人已登录。", steps(
        ("进入活动管理并填写总结标题、内容和图片", "表单校验通过"),
        ("点击发布总结", "总结保存成功"),
        ("打开活动详情", "总结内容展示在详情页")
    )),
    ("JG-REV-002", "签到评价", "参与者提交活动评分评价", "Priority 2", "覆盖 /activities/{id}/reviews。", "用户已报名或参与活动。", steps(
        ("进入活动详情评价区，选择 5 星并填写内容", "提交按钮可用"),
        ("点击提交评价", "评价成功"),
        ("刷新活动详情", "评价内容被展示或评分汇总更新")
    )),
    ("JG-AI-001", "AI策划", "AI 生成活动策划成功", "Priority 2", "覆盖 POST /ai/plans。", "打开 AI 策划页。", steps(
        ("输入主题 城市夜跑，填写人群和风格", "生成按钮可用"),
        ("点击生成", "返回活动名称、简介、建议时间、人数等方案"),
        ("将方案带入创建表单", "表单字段可编辑")
    )),
    ("JG-AI-002", "AI策划", "AI 主题为空时生成失败", "Priority 2", "覆盖 @NotBlank theme。", "打开 AI 策划页。", steps(
        ("主题留空点击生成", "系统阻止提交或返回活动主题不能为空"),
        ("补充主题后重试", "可正常生成")
    )),
    ("JG-AI-003", "AI策划", "AI 服务未配置时走降级方案", "Priority 2", "覆盖 IntegrationService 降级日志。", "不配置 QUJU_AI_API_KEY。", steps(
        ("输入合法主题并点击生成", "页面不崩溃"),
        ("查看生成结果或提示", "展示降级策划内容或明确失败提示，允许手动创建")
    )),
    ("JG-MAP-001", "地图地点", "地点搜索返回高德或降级结果", "Priority 2", "覆盖 GET /map/places。", "地图页或创建页可输入地点。", steps(
        ("搜索关键词 西湖，城市 杭州", "返回地点列表"),
        ("选择一个地点", "页面回填名称、地址和坐标")
    )),
    ("JG-FILE-001", "文件上传", "活动封面上传成功", "Priority 2", "覆盖 /files/upload 图片用途。", "普通用户已登录，准备 jpg/png 图片。", steps(
        ("在创建活动页上传封面", "上传成功并显示预览"),
        ("提交活动", "活动详情展示该封面")
    )),
    ("JG-TEAM-001", "兴趣小队", "查看小队列表", "Priority 1", "覆盖 GET /teams。", "后端有 team-01 等种子小队。", steps(
        ("进入小队页", "展示晚风散步社、周末逃跑计划等小队"),
        ("输入关键词 散步 搜索", "仅展示匹配小队")
    )),
    ("JG-TEAM-002", "兴趣小队", "创建公开加入小队", "Priority 1", "覆盖 POST /teams。", "普通用户已登录。", steps(
        ("填写小队名称、简介、标签、容量、加入方式公开加入", "表单校验通过"),
        ("点击创建", "创建成功，小队出现在列表中"),
        ("查看我的小队", "当前用户成为队长或成员")
    )),
    ("JG-TEAM-003", "兴趣小队", "加入公开小队成功", "Priority 1", "覆盖 POST /teams/{id}/members。", "用户未加入 team-01。", steps(
        ("打开公开小队详情", "显示加入按钮"),
        ("点击加入", "加入成功，成员数增加"),
        ("查看我的小队", "小队 id 出现在 memberships/me")
    )),
    ("JG-TEAM-004", "兴趣小队", "审核加入小队产生申请", "Priority 2", "覆盖审核加入模式。", "用户未加入 team-02。", steps(
        ("打开审核加入小队并点击申请加入", "申请提交成功"),
        ("队长查看加入申请", "申请出现在 join-requests 列表"),
        ("队长通过申请", "申请人成为小队成员")
    )),
    ("JG-TEAM-005", "兴趣小队", "队长拒绝加入申请", "Priority 2", "覆盖 rejectJoin。", "小队存在待审核申请。", steps(
        ("队长打开加入申请列表", "显示申请人信息"),
        ("填写拒绝原因并拒绝", "申请被拒绝"),
        ("申请人刷新页面", "未加入该小队，可看到失败状态或不在成员列表")
    )),
    ("JG-TEAM-006", "兴趣小队", "发布小队公告", "Priority 2", "覆盖 /announcements。", "队长已登录。", steps(
        ("进入小队管理，填写公告标题和内容", "公告可提交"),
        ("点击发布", "公告发布成功"),
        ("成员进入小队页", "可查看最新公告")
    )),
    ("JG-TEAM-007", "兴趣小队", "发起小队投票", "Priority 2", "覆盖 /polls。", "队长或成员已登录。", steps(
        ("填写投票主题和多个选项", "投票表单校验通过"),
        ("提交投票", "投票创建成功"),
        ("成员打开小队页", "可看到投票内容")
    )),
    ("JG-TEAM-008", "兴趣小队", "小队排行榜展示活跃成员", "Priority 3", "覆盖 /leaderboard。", "小队存在成员活动数据。", steps(
        ("打开小队详情排行榜", "展示成员排名、活跃度或贡献数据"),
        ("切换小队", "排行榜随小队 id 变化")
    )),
    ("JG-TEAM-009", "兴趣小队", "队长解散小队", "Priority 2", "覆盖 /dissolve。", "使用测试小队且当前用户为队长。", steps(
        ("点击解散小队并确认", "系统提示解散成功"),
        ("返回小队列表", "该小队不再作为正常小队展示")
    )),
    ("JG-SOC-001", "好友关注", "发送好友申请", "Priority 2", "覆盖 POST /friends/requests。", "普通用户已登录，目标用户存在。", steps(
        ("在用户资料页点击添加好友", "申请提交成功"),
        ("目标用户查看好友申请", "申请出现在列表中")
    )),
    ("JG-SOC-002", "好友关注", "同意好友申请", "Priority 2", "覆盖 approve。", "存在待处理好友申请。", steps(
        ("目标用户点击同意", "申请状态变为已通过"),
        ("双方查看好友列表", "对方出现在好友列表中")
    )),
    ("JG-SOC-003", "好友关注", "拒绝好友申请", "Priority 2", "覆盖 reject。", "存在待处理好友申请。", steps(
        ("目标用户点击拒绝", "申请状态变为已拒绝"),
        ("申请人查看好友列表", "对方不出现在好友列表中")
    )),
    ("JG-SOC-004", "好友关注", "更新好友备注和分组", "Priority 3", "覆盖 PUT /friends/{id}。", "双方已是好友。", steps(
        ("打开好友设置，修改备注和标签", "保存成功"),
        ("返回好友列表", "展示新的备注或分组")
    )),
    ("JG-SOC-005", "好友关注", "关注用户后粉丝数变化", "Priority 2", "覆盖 POST /follows/{id}。", "普通用户已登录。", steps(
        ("打开其他用户主页并点击关注", "关注成功"),
        ("刷新用户主页", "关注状态变为已关注，粉丝数增加")
    )),
    ("JG-SOC-006", "好友关注", "取消关注恢复状态", "Priority 2", "覆盖 DELETE /follows/{id}。", "已关注某用户。", steps(
        ("点击取消关注", "取消成功"),
        ("刷新用户主页", "按钮恢复为关注，粉丝数减少或保持一致")
    )),
    ("JG-SOC-007", "好友关注", "拉黑用户后不能互动", "Priority 2", "覆盖 /blocks/{id}。", "普通用户已登录。", steps(
        ("打开目标用户资料并点击拉黑", "系统提示拉黑成功"),
        ("尝试发送好友申请或消息", "系统阻止互动或不展示入口")
    )),
    ("JG-MSG-001", "消息聊天", "会话列表展示未读数和最后消息", "Priority 1", "覆盖 GET /conversations。", "普通用户已登录，存在种子会话。", steps(
        ("进入消息页", "展示会话名称、头像、最后消息和未读数"),
        ("点击一个会话", "进入消息详情")
    )),
    ("JG-MSG-002", "消息聊天", "查看会话消息历史", "Priority 1", "覆盖 GET /conversations/{id}/messages。", "普通用户已登录。", steps(
        ("打开 cv-01 会话", "加载历史消息"),
        ("滚动消息列表", "消息发送方、时间和内容展示正确")
    )),
    ("JG-MSG-003", "消息聊天", "发送文本消息成功", "Priority 1", "覆盖 POST /messages。", "普通用户已登录，存在会话。", steps(
        ("在输入框输入 明天见", "发送按钮可用"),
        ("点击发送", "消息出现在列表底部"),
        ("刷新会话", "新消息仍然存在")
    )),
    ("JG-MSG-004", "消息聊天", "标记消息已读", "Priority 2", "覆盖 /messages/{id}/read。", "存在未读消息。", steps(
        ("打开含未读消息的会话", "系统调用已读接口"),
        ("返回会话列表", "未读数减少或清零")
    )),
    ("JG-MSG-005", "消息聊天", "撤回本人消息", "Priority 2", "覆盖 /messages/{id}/recall。", "当前用户刚发送一条消息。", steps(
        ("点击本人消息的撤回操作", "撤回成功"),
        ("查看消息列表", "原内容被撤回提示替代或不再展示")
    )),
    ("JG-MSG-006", "消息聊天", "转发消息到其他会话", "Priority 3", "覆盖 /messages/{id}/forward。", "存在至少两个会话。", steps(
        ("选择一条消息并点击转发", "显示会话选择"),
        ("选择目标会话并确认", "目标会话出现转发消息")
    )),
    ("JG-MSG-007", "消息聊天", "未登录访问消息页被拦截", "Priority 1", "覆盖 requireToken。", "清空本地 token。", steps(
        ("直接访问 /messages", "系统跳转登录或提示请先登录"),
        ("调用会话接口", "返回 401 或统一未登录错误")
    )),
    ("JG-ADM-001", "运营后台", "管理员登录进入后台", "Priority 1", "覆盖管理员权限。", "存在 admin@quju.cn / Admin123456。", steps(
        ("访问 /admin 并输入管理员账号", "登录成功"),
        ("打开后台首页", "展示数据概览卡片")
    )),
    ("JG-ADM-002", "运营后台", "普通用户不能访问后台接口", "Priority 1", "覆盖 requireAdmin。", "普通用户已登录。", steps(
        ("访问 /admin/dashboard", "系统拒绝访问"),
        ("查看页面", "提示无权限或跳转用户首页")
    )),
    ("JG-ADM-003", "运营后台", "后台审核列表按类型筛选", "Priority 1", "覆盖 GET /admin/reviews。", "管理员已登录，存在种子审核任务。", steps(
        ("进入审核管理页", "展示活动审核、商家认证、举报处理等任务"),
        ("选择活动审核类型", "列表仅展示活动审核任务"),
        ("输入关键词搜索", "列表按关键词过滤")
    )),
    ("JG-ADM-004", "运营后台", "管理员通过审核任务", "Priority 1", "覆盖 approveReview。", "存在待审核任务。", steps(
        ("打开待审核活动任务", "可查看完整信息和风险原因"),
        ("点击通过", "任务状态变为已通过"),
        ("查看活动状态", "对应活动进入可展示或后续状态")
    )),
    ("JG-ADM-005", "运营后台", "管理员驳回审核任务必须填写原因", "Priority 1", "覆盖 rejectReview。", "存在待审核任务。", steps(
        ("点击驳回但不填写原因", "系统阻止提交或提示原因必填"),
        ("填写原因后驳回", "任务状态变为已驳回并记录原因")
    )),
    ("JG-ADM-006", "运营后台", "管理员要求修改审核任务", "Priority 2", "覆盖 requireChanges。", "存在待审核任务。", steps(
        ("填写修改意见并点击要求修改", "任务状态变为要求修改"),
        ("发起人查看活动", "可看到修改原因")
    )),
    ("JG-ADM-007", "运营后台", "用户列表按角色和状态筛选", "Priority 1", "覆盖 GET /admin/users。", "管理员已登录。", steps(
        ("进入用户管理页", "展示用户列表"),
        ("选择商家用户角色", "仅展示商家用户"),
        ("选择已封禁状态", "展示 momo77 等封禁用户")
    )),
    ("JG-ADM-008", "运营后台", "封禁用户并记录原因期限", "Priority 1", "覆盖 /users/{id}/ban。", "管理员已登录，目标用户正常。", steps(
        ("选择正常用户点击封禁", "弹出原因和期限表单"),
        ("填写原因和期限后确认", "用户状态变为已封禁"),
        ("目标用户尝试登录", "登录被拒绝并展示封禁信息")
    )),
    ("JG-ADM-009", "运营后台", "解封用户恢复登录", "Priority 1", "覆盖 /users/{id}/unblock。", "管理员已登录，存在封禁用户。", steps(
        ("点击解封", "用户状态恢复正常"),
        ("目标用户重新登录", "登录成功")
    )),
    ("JG-ADM-010", "运营后台", "活动下架和恢复", "Priority 1", "覆盖 /admin/activities offline/restore。", "管理员已登录，存在正常活动。", steps(
        ("在活动管理中选择活动并填写下架原因", "活动状态变为已下架"),
        ("用户端搜索该活动", "下架活动不再正常展示或显示不可报名"),
        ("管理员点击恢复", "活动恢复为可展示状态")
    )),
    ("JG-ADM-011", "运营后台", "小队停用和恢复", "Priority 2", "覆盖 /admin/teams stop/restore。", "管理员已登录，存在正常小队。", steps(
        ("选择小队并填写停用原因", "小队状态变为停用"),
        ("用户端查看小队列表", "停用小队不可加入或不展示"),
        ("管理员恢复小队", "小队恢复正常")
    )),
    ("JG-SEC-001", "权限异常", "未登录创建活动被拒绝", "Priority 1", "覆盖 Authorization 缺失。", "清空本地 token。", steps(
        ("直接调用创建活动接口或提交创建表单", "系统返回请先登录"),
        ("页面处理错误", "跳转登录或展示未登录提示")
    )),
    ("JG-SEC-002", "权限异常", "伪造 Token 访问接口被拒绝", "Priority 1", "覆盖 requireToken。", "浏览器或接口工具可设置请求头。", steps(
        ("设置 Authorization: Bearer invalid-token", "接口返回登录已过期或无效 token"),
        ("检查前端状态", "本地会话被清理")
    )),
    ("JG-SEC-003", "权限异常", "跨域请求允许本地前端访问后端", "Priority 2", "覆盖 CorsConfig。", "前端运行在 localhost:5173，后端在 localhost:8080。", steps(
        ("从前端页面发起 /api/activities 请求", "浏览器无 CORS 拦截"),
        ("查看网络面板", "接口响应正常返回")
    )),
    ("JG-SEC-004", "兼容性", "移动端主要页面自适应", "Priority 3", "覆盖 Vue 页面响应式布局。", "使用浏览器移动端模拟器。", steps(
        ("切换 iPhone 尺寸打开首页", "活动卡片、导航和筛选区不重叠"),
        ("打开创建页和详情页", "表单和按钮可正常操作")
    )),
    ("JG-SEC-005", "可用性", "后端服务异常时前端显示错误提示", "Priority 2", "覆盖 api.ts unwrap 异常处理。", "临时停止后端服务。", steps(
        ("刷新首页或提交登录", "请求失败被捕获"),
        ("查看页面", "显示友好错误提示，不出现空白页")
    )),
]


api_cases = [
    ("API-AUTH-001_普通用户登录成功", "验证 demo 用户登录返回 token", "post", "/auth/login", '{"email":"demo@quju.cn","password":"12345678"}', "包含", "token"),
    ("API-AUTH-002_管理员登录成功", "验证管理员登录返回用户信息", "post", "/auth/login", '{"email":"admin@quju.cn","password":"Admin123456"}', "包含", "token"),
    ("API-AUTH-003_错误密码登录失败", "验证错误密码不返回成功 token", "post", "/auth/login", '{"email":"demo@quju.cn","password":"bad-password"}', "包含", "密码"),
    ("API-AUTH-004_注册个人用户", "验证注册接口返回 activationToken", "post", "/auth/register", '{"email":"api_user_20260630@quju.cn","password":"12345678","confirmPassword":"12345678","nickname":"接口测试用户","role":"个人用户"}', "包含", "activationToken"),
    ("API-ACT-001_活动列表查询", "验证活动列表接口可返回种子活动", "get", "/activities", "", "包含", "落日以后"),
    ("API-ACT-002_关键词搜索桌游", "验证 keyword 参数过滤活动", "get", "/activities?keyword=桌游", "", "包含", "桌游"),
    ("API-ACT-003_分类筛选户外运动", "验证 category 参数", "get", "/activities?category=户外运动", "", "包含", "户外"),
    ("API-ACT-004_活动详情查询", "验证 act-001 详情", "get", "/activities/act-001", "", "包含", "运河"),
    ("API-ACT-005_活动详情不存在", "验证不存在 id 返回 404 或失败", "get", "/activities/not-exist", "", "包含", "404"),
    ("API-ACT-006_创建活动缺登录失败", "验证创建活动需要 Authorization", "post", "/activities", '{"title":"接口测试活动","summary":"接口测试","category":"学习交流","date":"2026-07-10","time":"19:00 - 21:00","location":"测试地点","capacity":8}', "包含", "登录"),
    ("API-DRAFT-001_保存草稿缺登录失败", "验证草稿保存需要 token", "post", "/activities/drafts", '{"title":"接口草稿"}', "包含", "登录"),
    ("API-REG-001_我的报名缺登录失败", "验证报名状态接口需要 token", "get", "/activities/registrations/me", "", "包含", "登录"),
    ("API-REG-002_报名缺登录失败", "验证报名接口需要 token", "post", "/activities/act-001/registrations", '{"fields":{"真实姓名":"测试用户","手机号码":"13800000000"}}', "包含", "登录"),
    ("API-CHK-001_扫码签到缺登录失败", "验证签到扫描需要 token", "post", "/activities/checkins/scan", '{"code":"invalid-code"}', "包含", "登录"),
    ("API-AI-001_AI策划生成", "验证 AI 策划接口返回方案或降级内容", "post", "/ai/plans", '{"theme":"城市夜跑","people":"大学生","style":"轻松"}', "包含", "title"),
    ("API-AI-002_AI策划主题为空", "验证 theme 必填校验", "post", "/ai/plans", '{"theme":"","people":"大学生","style":"轻松"}', "包含", "活动主题不能为空"),
    ("API-MAP-001_地点搜索", "验证地图地点搜索接口", "get", "/map/places?keyword=西湖&city=杭州", "", "包含", "西湖"),
    ("API-TEAM-001_小队列表", "验证小队列表返回种子小队", "get", "/teams", "", "包含", "晚风散步社"),
    ("API-TEAM-002_小队搜索", "验证小队 query 参数", "get", "/teams?query=散步", "", "包含", "散步"),
    ("API-TEAM-003_创建小队缺登录失败", "验证创建小队需要 token", "post", "/teams", '{"name":"接口小队","description":"接口测试","tags":["测试"],"capacity":20,"joinMode":"公开加入"}', "包含", "登录"),
    ("API-SOC-001_好友列表缺登录失败", "验证好友列表需要 token", "get", "/friends", "", "包含", "登录"),
    ("API-SOC-002_关注缺登录失败", "验证关注接口需要 token", "post", "/follows/u-003", "{}", "包含", "登录"),
    ("API-MSG-001_会话列表缺登录失败", "验证会话列表需要 token", "get", "/conversations", "", "包含", "登录"),
    ("API-MSG-002_发送消息缺登录失败", "验证发送消息需要 token", "post", "/conversations/cv-01/messages", '{"content":"hello"}', "包含", "登录"),
    ("API-ADMIN-001_后台概览缺登录失败", "验证后台接口需要管理员 token", "get", "/admin/dashboard", "", "包含", "登录"),
    ("API-ADMIN-002_审核列表缺登录失败", "验证审核列表权限", "get", "/admin/reviews", "", "包含", "登录"),
    ("API-ADMIN-003_用户列表缺登录失败", "验证用户列表权限", "get", "/admin/users", "", "包含", "登录"),
    ("API-ADMIN-004_活动下架缺登录失败", "验证活动下架权限", "post", "/admin/activities/act-001/offline", '{"reason":"接口测试"}', "包含", "登录"),
    ("API-ADMIN-005_小队停用缺登录失败", "验证小队停用权限", "post", "/admin/teams/team-01/stop", '{"reason":"接口测试"}', "包含", "登录"),
    ("API-FILE-001_文件上传缺登录失败", "验证上传接口需要 token", "post", "/files/upload", "{}", "包含", "登录"),
]


def build_manual():
    wb = load_workbook(MANUAL_TEMPLATE)
    ws = wb["sheet1"]
    max_col = 32
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for idx, case in enumerate(manual_cases, start=2):
        case_id, module, title, priority, desc, pre, step_values = case
        values = [
            f"{module}_{title}",
            case_id,
            "",
            "",
            "功能性测试",
            "手工测试",
            priority,
            "",
            "",
            "",
            "",
            desc,
            pre,
            "其他",
            module,
            "",
            "",
            "",
            *step_values,
        ]
        clone_row_style(ws, 2 if idx == 2 else idx - 1, idx, max_col)
        for col, value in enumerate(values, start=1):
            ws.cell(idx, col).value = value
    ws.freeze_panes = "A2"
    for col in range(1, max_col + 1):
        letter = ws.cell(1, col).column_letter
        if col == 1:
            ws.column_dimensions[letter].width = 44
        elif col in (12, 13, 19, 20, 21, 22, 23, 24, 25, 26):
            ws.column_dimensions[letter].width = 28
        elif col in (2, 5, 6, 7, 14, 15):
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 12
    wb.save(MANUAL_OUT)


def build_api():
    wb = load_workbook(API_TEMPLATE)
    ws = wb["Sheet1"]
    max_col = 11
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    base_url = "http://localhost:8080/api"
    for idx, (name, desc, method, path, body, match, expected) in enumerate(api_cases, start=2):
        values = [
            name,
            desc,
            method,
            "Content-Type=application/json",
            f"{base_url}{path}",
            "local",
            "qujuApi",
            "json" if method.lower() in ("post", "put", "delete") else "",
            body,
            match,
            expected,
        ]
        clone_row_style(ws, 2 if idx == 2 else idx - 1, idx, max_col)
        for col, value in enumerate(values, start=1):
            ws.cell(idx, col).value = value
    ws.freeze_panes = "A2"
    widths = [32, 34, 10, 28, 70, 14, 14, 14, 52, 18, 24]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width
    wb.save(API_OUT)


if __name__ == "__main__":
    build_manual()
    build_api()
    print(MANUAL_OUT)
    print(API_OUT)
    print(f"manual_cases={len(manual_cases)}")
    print(f"api_cases={len(api_cases)}")
